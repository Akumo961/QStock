"""
Reports endpoint — generates real downloadable files (PDF / Excel / CSV)
for the admin Reports page.

Report types (match frontend/src/components/admin/Reports.tsx exactly):
  - user_activity       : per-user borrow/return activity in a date range
  - inventory_status    : current snapshot of every item
  - transaction_history : full transaction list in a date range
  - overdue_items       : currently overdue borrows
  - usage_analytics     : summary stats + category breakdown in a date range
  - popular_items       : items ranked by borrow count in a date range
"""

import csv
import io
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from src.core.database import get_db
from src.core.security import get_current_admin_user
from src.models.item import Item
from src.models.review import Review
from src.models.transaction import Transaction, TransactionStatus
from src.models.user import User

router = APIRouter(tags=["Reports"])

VALID_REPORT_TYPES = {
    "user_activity",
    "inventory_status",
    "transaction_history",
    "overdue_items",
    "usage_analytics",
    "popular_items",
}
VALID_FORMATS = {"pdf", "excel", "csv"}


def _parse_date_range(start_date: Optional[str], end_date: Optional[str]):
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d") if start_date else datetime.utcnow() - timedelta(days=30)
        end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) if end_date else datetime.utcnow()
    except ValueError:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Dates must be in YYYY-MM-DD format")
    if start > end:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="start_date must be before end_date")
    return start, end


# ---------------------------------------------------------------------------
# Data builders — each returns (headers: list[str], rows: list[list[str]])
# ---------------------------------------------------------------------------

def _build_user_activity(db: Session, start, end):
    headers = ["User", "Email", "Department", "Borrows", "Returns", "Currently Borrowed", "Overdue"]
    users = db.query(User).all()
    rows = []
    for u in users:
        txns = db.query(Transaction).filter(
            Transaction.user_id == u.id,
            Transaction.borrowed_at >= start,
            Transaction.borrowed_at <= end,
        ).all()
        borrows = len(txns)
        returns = sum(1 for t in txns if t.status == TransactionStatus.RETURNED)
        active = sum(1 for t in txns if t.status == TransactionStatus.BORROWED)
        now = datetime.now(timezone.utc)
        overdue = sum(
            1 for t in txns
            if t.status == TransactionStatus.BORROWED and t.due_date and t.due_date < now
        )
        if borrows == 0:
            continue
        rows.append([u.full_name, u.email, u.department or "-", str(borrows), str(returns), str(active), str(overdue)])
    return headers, rows


def _build_inventory_status(db: Session, start, end):
    headers = ["Item Code", "Name", "Category", "Status", "Quantity", "Available", "Location"]
    items = db.query(Item).order_by(Item.name).all()
    rows = [
        [i.item_code, i.name, i.category.value if hasattr(i.category, "value") else str(i.category),
         i.status.value if hasattr(i.status, "value") else str(i.status),
         str(i.quantity), str(i.available_quantity), i.location or "-"]
        for i in items
    ]
    return headers, rows


def _build_transaction_history(db: Session, start, end):
    headers = ["Item", "User", "Status", "Quantity", "Borrowed At", "Due Date", "Returned At"]
    txns = db.query(Transaction).filter(
        Transaction.borrowed_at >= start, Transaction.borrowed_at <= end
    ).order_by(Transaction.borrowed_at.desc()).all()
    rows = []
    for t in txns:
        item = db.query(Item).filter(Item.id == t.item_id).first()
        user = db.query(User).filter(User.id == t.user_id).first()
        rows.append([
            item.name if item else f"Item #{t.item_id}",
            user.full_name if user else f"User #{t.user_id}",
            t.status.value if hasattr(t.status, "value") else str(t.status),
            str(t.quantity),
            t.borrowed_at.strftime("%Y-%m-%d %H:%M") if t.borrowed_at else "-",
            t.due_date.strftime("%Y-%m-%d") if t.due_date else "-",
            t.returned_at.strftime("%Y-%m-%d %H:%M") if t.returned_at else "-",
        ])
    return headers, rows


def _build_overdue_items(db: Session, start, end):
    headers = ["Item", "User", "Borrowed At", "Due Date", "Days Overdue"]
    now = datetime.now(timezone.utc)
    txns = db.query(Transaction).filter(
        Transaction.status == TransactionStatus.BORROWED,
        Transaction.due_date < now,
    ).order_by(Transaction.due_date).all()
    rows = []
    for t in txns:
        item = db.query(Item).filter(Item.id == t.item_id).first()
        user = db.query(User).filter(User.id == t.user_id).first()
        days_overdue = (now - t.due_date).days if t.due_date else 0
        rows.append([
            item.name if item else f"Item #{t.item_id}",
            user.full_name if user else f"User #{t.user_id}",
            t.borrowed_at.strftime("%Y-%m-%d") if t.borrowed_at else "-",
            t.due_date.strftime("%Y-%m-%d") if t.due_date else "-",
            str(days_overdue),
        ])
    return headers, rows


def _build_usage_analytics(db: Session, start, end):
    headers = ["Metric", "Value"]
    total_txns = db.query(Transaction).filter(Transaction.borrowed_at >= start, Transaction.borrowed_at <= end).count()
    total_returns = db.query(Transaction).filter(
        Transaction.borrowed_at >= start, Transaction.borrowed_at <= end,
        Transaction.status == TransactionStatus.RETURNED,
    ).count()
    total_reviews = db.query(Review).filter(Review.created_at >= start, Review.created_at <= end).count()
    issue_reviews = db.query(Review).filter(
        Review.created_at >= start, Review.created_at <= end, Review.has_issue.is_(True)
    ).count()
    return_rate = round((total_returns / total_txns * 100), 1) if total_txns else 0.0
    issue_rate = round((issue_reviews / total_reviews * 100), 1) if total_reviews else 0.0

    rows = [
        ["Total Borrows", str(total_txns)],
        ["Total Returns", str(total_returns)],
        ["Return Rate (%)", str(return_rate)],
        ["Total Reviews", str(total_reviews)],
        ["Issue Reports", str(issue_reviews)],
        ["Issue Rate (%)", str(issue_rate)],
    ]

    category_counts = (
        db.query(Item.category, func.count(Transaction.id))
        .join(Transaction, Transaction.item_id == Item.id)
        .filter(Transaction.borrowed_at >= start, Transaction.borrowed_at <= end)
        .group_by(Item.category)
        .all()
    )
    for category, count in category_counts:
        cat_label = category.value if hasattr(category, "value") else str(category)
        rows.append([f"Borrows in '{cat_label}'", str(count)])

    return headers, rows


def _build_popular_items(db: Session, start, end):
    headers = ["Item", "Category", "Borrow Count"]
    results = (
        db.query(Item.name, Item.category, func.count(Transaction.id).label("cnt"))
        .join(Transaction, Transaction.item_id == Item.id)
        .filter(Transaction.borrowed_at >= start, Transaction.borrowed_at <= end)
        .group_by(Item.id, Item.name, Item.category)
        .order_by(func.count(Transaction.id).desc())
        .limit(50)
        .all()
    )
    rows = [
        [name, category.value if hasattr(category, "value") else str(category), str(cnt)]
        for name, category, cnt in results
    ]
    return headers, rows


_BUILDERS = {
    "user_activity": _build_user_activity,
    "inventory_status": _build_inventory_status,
    "transaction_history": _build_transaction_history,
    "overdue_items": _build_overdue_items,
    "usage_analytics": _build_usage_analytics,
    "popular_items": _build_popular_items,
}


# ---------------------------------------------------------------------------
# Renderers — turn (headers, rows) into bytes for each format
# ---------------------------------------------------------------------------

def _render_csv(headers, rows) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _render_excel(headers, rows, title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows]) if rows else len(str(header))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_pdf(headers, rows, title: str, date_range_label: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles["Title"]),
        Paragraph(date_range_label, styles["Normal"]),
        Spacer(1, 0.3 * inch),
    ]

    if not rows:
        elements.append(Paragraph("No data found for the selected period.", styles["Normal"]))
    else:
        table_data = [headers] + [[str(c) for c in row] for row in rows]
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4332")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F8F4")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(table)

    doc.build(elements)
    return buf.getvalue()


_CONTENT_TYPES = {
    "csv": "text/csv",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}
_EXTENSIONS = {"csv": "csv", "excel": "xlsx", "pdf": "pdf"}


@router.get("/generate")
async def generate_report(
        report_type: str = Query(...),
        format: str = Query(...),
        start_date: Optional[str] = Query(None),
        end_date: Optional[str] = Query(None),
        db: Session = Depends(get_db),
        current_admin: User = Depends(get_current_admin_user),
):
    """Generate and download a report file (admin only)."""
    if report_type not in VALID_REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid report_type. Must be one of: {', '.join(sorted(VALID_REPORT_TYPES))}",
        )
    fmt = format.lower()
    if fmt not in VALID_FORMATS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid format. Must be one of: {', '.join(sorted(VALID_FORMATS))}",
        )

    start, end = _parse_date_range(start_date, end_date)

    try:
        headers, rows = _BUILDERS[report_type](db, start, end)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to build report data: {exc}",
        )

    title = report_type.replace("_", " ").title()
    date_label = f"{start.strftime('%Y-%m-%d')} to {(end - timedelta(days=1)).strftime('%Y-%m-%d')}"

    try:
        if fmt == "csv":
            content = _render_csv(headers, rows)
        elif fmt == "excel":
            content = _render_excel(headers, rows, title)
        else:
            content = _render_pdf(headers, rows, title, date_label)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to render {fmt} report: {exc}",
        )

    filename = f"{report_type}_{datetime.utcnow().strftime('%Y%m%d')}.{_EXTENSIONS[fmt]}"
    return StreamingResponse(
        io.BytesIO(content),
        media_type=_CONTENT_TYPES[fmt],
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )